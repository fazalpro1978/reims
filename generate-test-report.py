"""
Generates: prive-reims-test-report.xlsx
  Sheet 1 — Automated Results   (read-only summary, all 72 tests)
  Sheet 2 — Manual UI Checklist (editable: Result + Remarks columns)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colours ──────────────────────────────────────────────────────────────────
GOLD        = "C9A84C"
DARK        = "1A1A1A"
GREEN_DARK  = "166534"
GREEN_LIGHT = "DCFCE7"
GREEN_MID   = "22C55E"
RED_DARK    = "991B1B"
RED_LIGHT   = "FEE2E2"
BLUE_DARK   = "0369A1"
BLUE_LIGHT  = "E0F2FE"
PURPLE_DARK = "6D28D9"
PURPLE_LT   = "EDE9FE"
AMBER_DARK  = "B45309"
AMBER_LIGHT = "FEF3C7"
PINK_DARK   = "BE185D"
PINK_LIGHT  = "FCE7F3"
HEADER_BG   = "F1F5F9"
ROW_ALT     = "F8FAFC"
WHITE       = "FFFFFF"
BORDER_CLR  = "E2E8F0"
MUTED       = "94A3B8"
TEXT_MAIN   = "0F172A"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=TEXT_MAIN, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def border(color=BORDER_CLR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

wb = Workbook()

# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — AUTOMATED RESULTS
# ═════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Automated Results"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A6"

# ── Cover banner ─────────────────────────────────────────────────────────────
ws1.merge_cells("A1:F1")
c = ws1["A1"]
c.value = "PRIVÉ GROUP REAL ESTATE — RE-IMS"
c.font = Font(name="Calibri", bold=True, size=14, color=GOLD)
c.fill = fill(DARK)
c.alignment = align("center")
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:F2")
c = ws1["A2"]
c.value = "Automated Test Suite Results  ·  09 June 2026"
c.font = Font(name="Calibri", size=10, color="AAAAAA")
c.fill = fill(DARK)
c.alignment = align("center")
ws1.row_dimensions[2].height = 20

# ── Summary KPI row ──────────────────────────────────────────────────────────
ws1.row_dimensions[3].height = 8   # spacer

def kpi(ws, col_start, label, value, txt_dark, bg):
    col = col_start
    for c in [col, col+1]:
        cell = ws.cell(row=4, column=c)
        cell.fill = fill(bg)
        cell.border = border(txt_dark)
    ws.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_start+1)
    c = ws.cell(row=4, column=col_start)
    c.value = f"{value}   {label}"
    c.font = Font(name="Calibri", bold=True, size=12, color=txt_dark)
    c.alignment = align("center")
    ws.row_dimensions[4].height = 28

kpi(ws1, 1, "TESTS PASSED", "✓  72", GREEN_DARK, GREEN_LIGHT)
kpi(ws1, 3, "TESTS FAILED", "✗   0", RED_DARK,   RED_LIGHT)
kpi(ws1, 5, "SUITES", "◆   5",  BLUE_DARK,  BLUE_LIGHT)

ws1.row_dimensions[5].height = 8  # spacer

# ── Column headers ────────────────────────────────────────────────────────────
headers = ["Suite", "Describe Group", "Test Name", "Result", "Time", "Status"]
col_widths = [30, 28, 60, 10, 8, 10]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws1.cell(row=6, column=i, value=h)
    c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    c.fill = fill(DARK)
    c.alignment = align("center")
    c.border = border(DARK)
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[6].height = 22

# ── Test data ─────────────────────────────────────────────────────────────────
tests = [
    # (suite, group, name, time)
    # legalDuration
    ("legalDuration.test.ts", "parseLegalDuration", 'parses "1 Year"', "2 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", 'parses "2 Years"', "<1 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", 'parses "6 Months"', "2 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", 'parses "3 Weeks"', "<1 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", 'parses "30 Days"', "<1 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", 'is case-insensitive ("12 MONTHS")', "<1 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", 'handles singular "day"', "1 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", 'handles decimal values ("1.5 years")', "<1 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", "returns default {1, Years} for empty string", "3 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", "returns default {1, Years} for unrecognised string", "1 ms"),
    ("legalDuration.test.ts", "parseLegalDuration", "trims leading/trailing whitespace", "<1 ms"),
    ("legalDuration.test.ts", "calcEndDate", "returns empty string when start is empty", "<1 ms"),
    ("legalDuration.test.ts", "calcEndDate", "adds 1 year correctly", "1 ms"),
    ("legalDuration.test.ts", "calcEndDate", "adds 6 months correctly", "<1 ms"),
    ("legalDuration.test.ts", "calcEndDate", "adds 2 weeks (14 days)", "<1 ms"),
    ("legalDuration.test.ts", "calcEndDate", "adds 30 days", "2 ms"),
    ("legalDuration.test.ts", "calcEndDate", "handles year boundary (Dec → Jan)", "<1 ms"),
    ("legalDuration.test.ts", "calcEndDate", "handles leap year correctly (Feb 28 + 1 day = Feb 29)", "<1 ms"),
    ("legalDuration.test.ts", "calcDurationFromDates", "returns {1, Years} when start is empty", "<1 ms"),
    ("legalDuration.test.ts", "calcDurationFromDates", "returns {1, Years} when end is empty", "<1 ms"),
    ("legalDuration.test.ts", "calcDurationFromDates", "returns {1, Days} when end is before start", "<1 ms"),
    ("legalDuration.test.ts", "calcDurationFromDates", "detects exact 1-year span", "<1 ms"),
    ("legalDuration.test.ts", "calcDurationFromDates", "detects exact 2-year span (no leap year)", "<1 ms"),
    ("legalDuration.test.ts", "calcDurationFromDates", "detects exact month span (30 days)", "2 ms"),
    ("legalDuration.test.ts", "calcDurationFromDates", "detects exact week span (14 days)", "<1 ms"),
    ("legalDuration.test.ts", "calcDurationFromDates", "returns days for non-divisible span (15 days)", "1 ms"),
    # shareUtils — note: we listed 18 but actual suite has 17 (4+7+6)
    ("shareUtils.test.ts", "formatQAR", "formats zero", "1 ms"),
    ("shareUtils.test.ts", "formatQAR", "formats thousands with comma separator", "<1 ms"),
    ("shareUtils.test.ts", "formatQAR", 'formats large rent value', "1 ms"),
    ("shareUtils.test.ts", "formatQAR", 'formats decimal value (begins with "QAR")', "<1 ms"),
    ("shareUtils.test.ts", "generateShareText", "includes property name", "<1 ms"),
    ("shareUtils.test.ts", "generateShareText", "includes unit number", "<1 ms"),
    ("shareUtils.test.ts", "generateShareText", "includes zone and zone code", "1 ms"),
    ("shareUtils.test.ts", "generateShareText", "includes rent amount", "<1 ms"),
    ("shareUtils.test.ts", "generateShareText", "includes MOCI code", "<1 ms"),
    ("shareUtils.test.ts", "generateShareText", "includes status without underscore (Under Maintenance)", "<1 ms"),
    ("shareUtils.test.ts", "generateShareText", "is a single line (no newlines)", "1 ms"),
    ("shareUtils.test.ts", "generateEmailBody", "includes header line (Property Listing Enquiry)", "<1 ms"),
    ("shareUtils.test.ts", "generateEmailBody", 'includes property details (name, unit no, config)', "<1 ms"),
    ("shareUtils.test.ts", "generateEmailBody", 'includes rent formatted with "/ month"', "1 ms"),
    ("shareUtils.test.ts", "generateEmailBody", "includes company contact details (+974 7707 5959)", "<1 ms"),
    ("shareUtils.test.ts", "generateEmailBody", "includes licence and CR numbers (773 / 187753)", "<1 ms"),
    ("shareUtils.test.ts", "generateEmailBody", "is a multi-line string (> 5 lines)", "<1 ms"),
    # auditLog
    ("auditLog.test.ts", "logEvent", "does nothing when unitId is empty", "1 ms"),
    ("auditLog.test.ts", "logEvent", "inserts a row with the correct shape", "1 ms"),
    ("auditLog.test.ts", "logEvent", "sets tab_context to null when tab is omitted", "1 ms"),
    ("auditLog.test.ts", "logEvent", "stores field, old/new values when provided", "<1 ms"),
    ("auditLog.test.ts", "logEvent", "stores payload when provided", "<1 ms"),
    ("auditLog.test.ts", "logEvent", "defaults payload to {} when omitted", "1 ms"),
    ("auditLog.test.ts", "logEvent", "is silent when Supabase insert throws", "<1 ms"),
    ("auditLog.test.ts", "logEvent", "covers all 10 AuditAction values without throwing", "1 ms"),
    # upload API
    ("api/upload.test.ts", "POST — parameter validation", "returns 400 when both file and path are missing", "8 ms"),
    ("api/upload.test.ts", "POST — parameter validation", "returns 400 when file is missing but path is present", "2 ms"),
    ("api/upload.test.ts", "POST — parameter validation", "returns 400 when path is missing but file is present", "5 ms"),
    ("api/upload.test.ts", "POST — file type validation", "accepts application/pdf → 200", "2 ms"),
    ("api/upload.test.ts", "POST — file type validation", "accepts image/jpeg → 200", "1 ms"),
    ("api/upload.test.ts", "POST — file type validation", "accepts image/png → 200", "2 ms"),
    ("api/upload.test.ts", "POST — file type validation", "accepts image/webp → 200", "1 ms"),
    ("api/upload.test.ts", "POST — file type validation", "rejects application/exe → 400", "1 ms"),
    ("api/upload.test.ts", "POST — file type validation", "rejects text/plain → 400", "1 ms"),
    ("api/upload.test.ts", "POST — file type validation", "rejects application/zip → 400", "1 ms"),
    ("api/upload.test.ts", "POST — file type validation", "rejects application/javascript → 400", "1 ms"),
    ("api/upload.test.ts", "POST — file size & Supabase error", "accepts file exactly at 50 MB limit → 200", "174 ms"),
    ("api/upload.test.ts", "POST — file size & Supabase error", "rejects file 1 byte over 50 MB → 400", "188 ms"),
    ("api/upload.test.ts", "POST — file size & Supabase error", "returns 500 when Supabase storage returns an error", "2 ms"),
    # signed-url API
    ("api/signed-url.test.ts", "GET — parameter & security", "returns 400 when path param is missing", "2 ms"),
    ('api/signed-url.test.ts', "GET — parameter & security", 'rejects path with ".." (path traversal attack)', "1 ms"),
    ("api/signed-url.test.ts", "GET — parameter & security", 'rejects absolute path starting with "/"', "1 ms"),
    ("api/signed-url.test.ts", "GET — parameter & security", "rejects embedded traversal (units/abc/../../../etc/passwd)", "<1 ms"),
    ("api/signed-url.test.ts", "GET — success & error", "returns 200 with signedUrl for a valid path", "1 ms"),
    ("api/signed-url.test.ts", "GET — success & error", "calls createSignedUrl with 3600-second (1-hour) expiry", "1 ms"),
    ("api/signed-url.test.ts", "GET — success & error", "returns 500 when Supabase returns an error", "<1 ms"),
]

SUITE_COLORS = {
    "legalDuration.test.ts": ("EFF6FF", "1D4ED8"),
    "shareUtils.test.ts":    ("F0FDF4", "166534"),
    "auditLog.test.ts":      ("FDF4FF", "7E22CE"),
    "api/upload.test.ts":    ("FFF7ED", "C2410C"),
    "api/signed-url.test.ts":("ECFDF5", "065F46"),
}

prev_suite = None
for idx, (suite, group, name, time) in enumerate(tests):
    row = idx + 7
    bg, txt = SUITE_COLORS.get(suite, (WHITE, TEXT_MAIN))
    row_bg = bg if idx % 2 == 0 else "FFFFFF"

    cells = [suite, group, name, "PASS", time, "✓"]
    for col, val in enumerate(cells, 1):
        c = ws1.cell(row=row, column=col, value=val)
        c.fill = fill(row_bg)
        c.border = border()
        c.font = Font(name="Calibri", size=9,
                      color=(GREEN_DARK if col == 6 else TEXT_MAIN))
        c.alignment = align("center" if col in (4, 5, 6) else "left",
                            wrap=(col == 3))
    ws1.row_dimensions[row].height = 16

# ── Totals footer ─────────────────────────────────────────────────────────────
footer_row = len(tests) + 7
ws1.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=5)
c = ws1.cell(row=footer_row, column=1,
             value=f"TOTAL  ·  {len(tests)} tests  ·  5 suites  ·  0.831 s  ·  Run: npm test")
c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
c.fill = fill(DARK)
c.alignment = align("center")
c.border = border(DARK)

c6 = ws1.cell(row=footer_row, column=6, value="ALL GREEN")
c6.font = Font(name="Calibri", bold=True, size=9, color=GREEN_LIGHT)
c6.fill = fill(GREEN_DARK)
c6.alignment = align("center")
c6.border = border(DARK)
ws1.row_dimensions[footer_row].height = 20


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — MANUAL UI CHECKLIST
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Manual UI Checklist")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A6"

# ── Cover banner ─────────────────────────────────────────────────────────────
ws2.merge_cells("A1:G1")
c = ws2["A1"]
c.value = "PRIVÉ GROUP REAL ESTATE — RE-IMS"
c.font = Font(name="Calibri", bold=True, size=14, color=GOLD)
c.fill = fill(DARK)
c.alignment = align("center")
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:G2")
c = ws2["A2"]
c.value = "Manual UI Test Checklist  ·  Tester: ___________________________  ·  Date: _______________"
c.font = Font(name="Calibri", size=10, color="AAAAAA")
c.fill = fill(DARK)
c.alignment = align("center")
ws2.row_dimensions[2].height = 20

ws2.row_dimensions[3].height = 8

# ── Instructions ─────────────────────────────────────────────────────────────
ws2.merge_cells("A4:G4")
c = ws2["A4"]
c.value = ("  ℹ  For each test: select PASS, FAIL, or SKIP from the Result dropdown.  "
           "Add notes in the Remarks column.  Highlighted rows are security-critical.")
c.font = Font(name="Calibri", size=9, color=AMBER_DARK, italic=True)
c.fill = fill(AMBER_LIGHT)
c.alignment = align("left")
ws2.row_dimensions[4].height = 18

ws2.row_dimensions[5].height = 8

# ── Column headers ────────────────────────────────────────────────────────────
col_headers = ["#", "Category", "Test Description", "Tag", "Priority", "Result", "Remarks"]
col_widths2  = [4,   22,         62,                  10,    9,          10,        38]
for i, (h, w) in enumerate(zip(col_headers, col_widths2), 1):
    c = ws2.cell(row=6, column=i, value=h)
    c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    c.fill = fill(DARK)
    c.alignment = align("center")
    c.border = border(DARK)
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[6].height = 22

# ── Add data validation for Result column (col 6) ────────────────────────────
dv = DataValidation(
    type="list",
    formula1='"PASS,FAIL,SKIP,BLOCKED"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid entry",
    error='Choose PASS, FAIL, SKIP, or BLOCKED'
)
ws2.add_data_validation(dv)

# ── Manual test data ──────────────────────────────────────────────────────────
# (category, description, tag, priority, is_security)
manual_tests = [
    # Authentication
    ("Auth & Access",    "Enter admin PIN PRIVE2024 → all fields unlock for editing",                                              "SECURITY", "P1", True),
    ("Auth & Access",    "Lock again (navigate away) → fields return to read-only mode",                                           "SECURITY", "P1", True),
    ("Auth & Access",    "Enter a wrong PIN → error shown, fields stay locked",                                                    "SECURITY", "P1", True),
    # Inventory
    ("Inventory Table",  "Search box filters rows in real-time as you type (property, unit no, zone)",                             "UX",       "P1", False),
    ("Inventory Table",  "Status filter dropdown shows/hides rows by status correctly",                                            "STATE",    "P1", False),
    ("Inventory Table",  "Furnishing filter works in combination with status filter (both active simultaneously)",                  "STATE",    "P1", False),
    ("Inventory Table",  "Pagination: Next/Prev buttons work; page number badge updates correctly",                                "UX",       "P2", False),
    ("Inventory Table",  "Context menu (right-click or ⋮) appears near the row without clipping viewport edge",                   "UX",       "P2", False),
    ("Inventory Table",  'Duplicate from context menu: creates -COPY row, table refreshes, success toast appears',                 "DB",       "P1", False),
    ("Inventory Table",  "Toast auto-dismisses after ~4–5 seconds",                                                               "UX",       "P3", False),
    # Modal general
    ("Modal",            "Clicking a row opens the slide-out modal with correct unit data pre-populated",                          "DB",       "P1", False),
    ("Modal",            "A RECORD_VIEW event appears in the Operational System Log after opening the modal",                      "DB",       "P1", False),
    ("Modal",            "Close button (×) closes the modal; clicking the backdrop also closes it",                                "UX",       "P2", False),
    ("Modal",            "Switching tabs logs a TAB_NAVIGATE event in the System Log with the correct tab name",                   "DB",       "P2", False),
    # Property tab
    ("Property Tab",     "Realtor dropdown selection auto-fills the MOCI code field",                                              "STATE",    "P1", False),
    ("Property Tab",     "Zone dropdown selection auto-fills the zone code field",                                                 "STATE",    "P1", False),
    ("Property Tab",     "Status dropdown change updates the status badge colour on save",                                         "STATE",    "P1", False),
    ("Property Tab",     "Amenities toggle buttons add/remove items; amenity count badge updates live",                            "STATE",    "P2", False),
    ("Property Tab",     "Save → close → reopen: all Property fields match what was saved",                                       "DB",       "P1", False),
    # Financials tab
    ("Financials Tab",   "Rent field updates the QAR x,xxx formatted display live",                                               "STATE",    "P1", False),
    ("Financials Tab",   "Agency fee toggle shows/hides fee amount and 'paid by' fields",                                         "STATE",    "P1", False),
    ("Financials Tab",   "Save → reload: all financial fields persist correctly",                                                  "DB",       "P1", False),
    # Commission — duration
    ("Commission — Duration", "Change start date → end date auto-calculates (start + duration)",                                  "STATE",    "P1", False),
    ("Commission — Duration", "Change duration value/unit → end date auto-recalculates",                                          "STATE",    "P1", False),
    ("Commission — Duration", "Change end date → duration auto-recalculates back to match",                                       "STATE",    "P1", False),
    ("Commission — Duration", "Client type toggle Individual ↔ Company: Authorized Signatory field appears/disappears",            "STATE",    "P1", False),
    ("Commission — Duration", "Property registration status: 'Registration by' field visible only for relevant statuses",          "STATE",    "P2", False),
    # Commission — documents
    ("Commission — Docs", "QID row: Upload a PDF → spinner → green done card with filename displayed",                            "DB",       "P1", False),
    ("Commission — Docs", "QID row: Close and reopen modal → file still shows in green done state (async prop sync)",             "DB",       "P1", False),
    ("Commission — Docs", "QID row: Click View button → signed URL opens document in new tab",                                    "SECURITY", "P1", True),
    ("Commission — Docs", "QID row: Click × remove → card resets to idle Upload file state",                                     "STATE",    "P1", False),
    ("Commission — Docs", "Passport row: Paste external URL → Save Link → blue saved state with clickable link",                  "DB",       "P1", False),
    ("Commission — Docs", "Passport row: Remove link → resets to URL input state",                                                "STATE",    "P2", False),
    ("Commission — Docs", "Other Documents row: Label input accepts custom text; label persists on save",                         "DB",       "P2", False),
    ("Commission — Docs", "Upload file AND save URL on same row → card turns green Verified; sections are independent",           "STATE",    "P1", False),
    ("Commission — Docs", "Commission Contract and Reg Certificate rows: same upload/URL/remove flows work",                      "DB",       "P1", False),
    ("Commission — Docs", "Commission & Legal documents do NOT appear in the public PDF report (privacy boundary)",               "SECURITY", "P1", True),
    # Operational tab
    ("Operational Tab",  "Inspection Report / Inventory Checklist / Handover Certificate: full upload, view, remove cycle",       "DB",       "P1", False),
    ("Operational Tab",  "Editing a text field then blurring triggers a FIELD_UPDATE event in the System Log",                   "DB",       "P1", False),
    ("Operational Tab",  "Save button triggers a RECORD_SAVE event in the System Log with payload",                              "DB",       "P1", False),
    # System log
    ("System Log",       "System Log shows events from all 4 tabs (RECORD_VIEW, TAB_NAVIGATE, FILE_UPLOAD, RECORD_SAVE etc.)",   "DB",       "P1", False),
    ("System Log",       "All 10 action type badges render with correct colours (RECORD_VIEW=purple, FILE_UPLOAD=blue etc.)",     "UX",       "P2", False),
    ("System Log",       "Each log entry shows: timestamp, operator (Administrator), tab context, and payload details",           "DB",       "P1", False),
    ("System Log",       "Refresh button re-loads events from DB (no stale cache)",                                              "STATE",    "P2", False),
    ("System Log",       "Entry count reads 'immutable events · complete record history' — no delete/clear button present",       "SECURITY", "P1", True),
    # PDF report
    ("PDF Report",       "Report page opens and all property metadata renders correctly (name, unit, zone, rent, type, config)",   "DB",       "P1", False),
    ("PDF Report",       "Print/Download button triggers browser print dialog",                                                    "UX",       "P2", False),
    ("PDF Report",       "No commission/legal document links visible in the report output",                                        "SECURITY", "P1", True),
    # Sharing
    ("Sharing",          "WhatsApp button opens wa.me with correct unit details URL-encoded in the message",                      "E2E",      "P2", False),
    ("Sharing",          "Email button opens mailto: with correct subject line and pre-filled body",                               "E2E",      "P2", False),
    ("Sharing",          "Copy button copies share text to clipboard; confirmation toast appears",                                 "UX",       "P2", False),
    # Theme
    ("Theme Switcher",   "Selecting a theme applies CSS variable changes immediately",                                             "UX",       "P3", False),
    ("Theme Switcher",   "Theme choice persisted in localStorage; page reload retains the selection",                              "STATE",    "P3", False),
    # Navigation
    ("Navigation",       "Side nav opens/closes via hamburger; Esc key also closes it",                                           "UX",       "P2", False),
    ("Navigation",       "Body scroll locks when side nav is open",                                                               "UX",       "P3", False),
    ("Navigation",       "Active route is highlighted in the nav",                                                                "UX",       "P3", False),
    ("Navigation",       "Mobile viewport (<640 px): floating action button visible, hamburger in header",                        "UX",       "P2", False),
    ("Navigation",       "Desktop viewport: full layout, correct spacing, no overflow clipping",                                  "UX",       "P2", False),
    # Error states
    ("Error States",     "DevTools → Offline → attempt upload → shows 'Network error — please retry' error state",               "E2E",      "P1", False),
    ("Error States",     "Retry button in error state returns the row to idle",                                                   "STATE",    "P1", False),
    ("Error States",     "If Supabase unreachable, inventory table shows DB error banner (not a blank screen)",                   "E2E",      "P2", False),
]

TAG_COLORS = {
    "SECURITY": (AMBER_DARK,  AMBER_LIGHT),
    "E2E":      (PURPLE_DARK, PURPLE_LT),
    "STATE":    (BLUE_DARK,   BLUE_LIGHT),
    "UX":       (PINK_DARK,   PINK_LIGHT),
    "DB":       (GREEN_DARK,  GREEN_LIGHT),
}

PRIO_COLORS = {
    "P1": ("991B1B", "FEE2E2"),
    "P2": ("1E3A5F", "DBEAFE"),
    "P3": ("374151", "F3F4F6"),
}

prev_cat = None
for idx, (cat, desc, tag, prio, is_sec) in enumerate(manual_tests):
    row = idx + 7
    row_bg = "FFF8EE" if is_sec else (ROW_ALT if idx % 2 == 0 else WHITE)

    # Col 1 — number
    c = ws2.cell(row=row, column=1, value=idx + 1)
    c.font = Font(name="Calibri", size=8, color=MUTED)
    c.fill = fill(row_bg)
    c.alignment = align("center")
    c.border = border()

    # Col 2 — category
    c = ws2.cell(row=row, column=2, value=cat)
    c.font = Font(name="Calibri", size=9, bold=(cat != prev_cat))
    c.fill = fill(row_bg)
    c.alignment = align("left", wrap=False)
    c.border = border()
    prev_cat = cat

    # Col 3 — description
    c = ws2.cell(row=row, column=3, value=desc)
    c.font = Font(name="Calibri", size=9)
    c.fill = fill(row_bg)
    c.alignment = align("left", wrap=True)
    c.border = border()

    # Col 4 — tag badge
    tag_txt, tag_bg = TAG_COLORS.get(tag, (TEXT_MAIN, WHITE))
    c = ws2.cell(row=row, column=4, value=tag)
    c.font = Font(name="Calibri", bold=True, size=8, color=tag_txt)
    c.fill = fill(tag_bg)
    c.alignment = align("center")
    c.border = border(tag_txt)

    # Col 5 — priority
    p_txt, p_bg = PRIO_COLORS.get(prio, (TEXT_MAIN, WHITE))
    c = ws2.cell(row=row, column=5, value=prio)
    c.font = Font(name="Calibri", bold=True, size=8, color=p_txt)
    c.fill = fill(p_bg)
    c.alignment = align("center")
    c.border = border(p_txt)

    # Col 6 — Result (dropdown)
    c = ws2.cell(row=row, column=6, value="")
    c.fill = fill("FAFAFA")
    c.alignment = align("center")
    c.border = border("888888")
    dv.add(c)

    # Col 7 — Remarks
    c = ws2.cell(row=row, column=7, value="")
    c.fill = fill("FAFAFA")
    c.alignment = align("left", wrap=True)
    c.border = border("888888")
    c.font = Font(name="Calibri", size=9, italic=True, color="64748B")

    ws2.row_dimensions[row].height = 18

# ── Conditional formatting: PASS=green, FAIL=red, SKIP=grey ──────────────────
result_range = f"F7:F{len(manual_tests) + 6}"

pass_fill = PatternFill("solid", bgColor="DCFCE7")
pass_font = Font(name="Calibri", bold=True, color="166534")
fail_fill = PatternFill("solid", bgColor="FEE2E2")
fail_font = Font(name="Calibri", bold=True, color="991B1B")
skip_fill = PatternFill("solid", bgColor="F1F5F9")
skip_font = Font(name="Calibri", bold=True, color="475569")
blok_fill = PatternFill("solid", bgColor="FFF7ED")
blok_font = Font(name="Calibri", bold=True, color="C2410C")

ws2.conditional_formatting.add(result_range, FormulaRule(
    formula=['F7="PASS"'],  stopIfTrue=True,
    fill=pass_fill, font=pass_font))
ws2.conditional_formatting.add(result_range, FormulaRule(
    formula=['F7="FAIL"'],  stopIfTrue=True,
    fill=fail_fill, font=fail_font))
ws2.conditional_formatting.add(result_range, FormulaRule(
    formula=['F7="SKIP"'],  stopIfTrue=True,
    fill=skip_fill, font=skip_font))
ws2.conditional_formatting.add(result_range, FormulaRule(
    formula=['F7="BLOCKED"'], stopIfTrue=True,
    fill=blok_fill, font=blok_font))

# ── Summary footer ────────────────────────────────────────────────────────────
footer = len(manual_tests) + 7
ws2.merge_cells(start_row=footer, start_column=1, end_row=footer, end_column=5)
total_count = len(manual_tests)
sec_count   = sum(1 for t in manual_tests if t[4])
p1_count    = sum(1 for t in manual_tests if t[3] == "P1")

c = ws2.cell(row=footer, column=1,
             value=(f"  {total_count} tests total  ·  "
                    f"{sec_count} security-critical (amber rows)  ·  "
                    f"{p1_count} P1 priority  ·  "
                    f"Result options: PASS / FAIL / SKIP / BLOCKED"))
c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
c.fill = fill(DARK)
c.alignment = align("left")
c.border = border(DARK)

pass_f  = f'=COUNTIF(F7:F{footer-1},"PASS")'
fail_f  = f'=COUNTIF(F7:F{footer-1},"FAIL")'

c_pass = ws2.cell(row=footer, column=6, value=pass_f)
c_pass.font = Font(name="Calibri", bold=True, size=9, color=GREEN_LIGHT)
c_pass.fill = fill(GREEN_DARK)
c_pass.alignment = align("center")
c_pass.border = border(DARK)

c_fail = ws2.cell(row=footer, column=7,
                  value=f'=COUNTIF(F7:F{footer-1},"FAIL")&" FAIL / "&COUNTIF(F7:F{footer-1},"PASS")&" PASS"')
c_fail.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
c_fail.fill = fill(DARK)
c_fail.alignment = align("center")
c_fail.border = border(DARK)

ws2.row_dimensions[footer].height = 20


# ═════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — LEGEND
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Legend")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:D1")
c = ws3["A1"]
c.value = "Legend & Colour Reference"
c.font = Font(name="Calibri", bold=True, size=13, color=GOLD)
c.fill = fill(DARK)
c.alignment = align("center")
ws3.row_dimensions[1].height = 28
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 24
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 30

legends = [
    ("RESULT VALUES", "", "", ""),
    ("PASS",    "Test passed successfully",           "Green",  "Select when test behaves as expected"),
    ("FAIL",    "Test produced incorrect behaviour",  "Red",    "Add details in Remarks — must be raised as a bug"),
    ("SKIP",    "Test not applicable / not run",      "Grey",   "Note reason in Remarks"),
    ("BLOCKED", "Test could not be run",              "Amber",  "Note blocker in Remarks"),
    ("", "", "", ""),
    ("TAGS", "", "", ""),
    ("SECURITY", "Privacy / access control check",   "Amber",  "These rows are highlighted — do not skip"),
    ("DB",       "Requires live Supabase connection", "Green",  "Run against reims-sigma.vercel.app"),
    ("STATE",    "UI state / reactivity check",       "Blue",   "Can be tested locally (npm run dev)"),
    ("UX",       "Visual / interaction quality",      "Pink",   "Subjective — use judgement"),
    ("E2E",      "End-to-end integration test",       "Purple", "Requires live environment"),
    ("", "", "", ""),
    ("PRIORITY", "", "", ""),
    ("P1", "Must pass before release", "Red",  "Blocking — address immediately if FAIL"),
    ("P2", "Should pass before release","Blue","Non-blocking but important"),
    ("P3", "Nice to have",             "Grey", "Can be deferred"),
]

row_bg_map = {
    "PASS": (GREEN_DARK, GREEN_LIGHT),
    "FAIL": (RED_DARK,   RED_LIGHT),
    "SKIP": ("475569",   "F1F5F9"),
    "BLOCKED": (AMBER_DARK, AMBER_LIGHT),
    "SECURITY": (AMBER_DARK, AMBER_LIGHT),
    "DB":       (GREEN_DARK, GREEN_LIGHT),
    "STATE":    (BLUE_DARK,  BLUE_LIGHT),
    "UX":       (PINK_DARK,  PINK_LIGHT),
    "E2E":      (PURPLE_DARK, PURPLE_LT),
    "P1": ("991B1B", "FEE2E2"),
    "P2": ("1E3A5F", "DBEAFE"),
    "P3": ("374151", "F3F4F6"),
}

section_labels = {"RESULT VALUES", "TAGS", "PRIORITY"}

for i, (key, desc, color, note) in enumerate(legends, 2):
    if not key:
        ws3.row_dimensions[i].height = 6
        continue
    if key in section_labels:
        ws3.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        c = ws3.cell(row=i, column=1, value=key)
        c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        c.fill = fill(DARK)
        c.alignment = align("center")
        ws3.row_dimensions[i].height = 18
        continue

    txt_c, bg_c = row_bg_map.get(key, (TEXT_MAIN, WHITE))
    for col in range(1, 5):
        c = ws3.cell(row=i, column=col)
        c.fill = fill(bg_c)
        c.border = border()
    ws3.cell(row=i, column=1, value=key).font = Font(name="Calibri", bold=True, size=9, color=txt_c)
    ws3.cell(row=i, column=1).alignment = align("center")
    ws3.cell(row=i, column=2, value=desc).font = Font(name="Calibri", size=9)
    ws3.cell(row=i, column=3, value=color).font = Font(name="Calibri", size=9, color=MUTED)
    ws3.cell(row=i, column=3).alignment = align("center")
    ws3.cell(row=i, column=4, value=note).font = Font(name="Calibri", size=9, italic=True)
    ws3.row_dimensions[i].height = 16


# ── Save ──────────────────────────────────────────────────────────────────────
path = "prive-reims-test-report.xlsx"
wb.save(path)
print(f"Saved: {path}")
